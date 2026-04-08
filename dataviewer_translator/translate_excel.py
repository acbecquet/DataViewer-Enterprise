"""
Excel Bidirectional Translator
Translates all cells in an Excel file between any two supported languages using Claude API.
"""

import os
import time
import re
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from anthropic import Anthropic

# Initialize Claude API client
# Set your API key as an environment variable: ANTHROPIC_API_KEY
client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

# Languages whose scripts are compact (CJK) — useful for font-size logic in other modules
CJK_LANGUAGES = {"Chinese (Simplified)", "Chinese (Traditional)", "Japanese", "Korean"}


def select_file():
    """Open a file dialog to select an Excel file."""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    file_path = filedialog.askopenfilename(
        title="Select Excel file to translate",
        filetypes=[
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        ]
    )

    root.destroy()
    return file_path if file_path else None


def sanitize_filename(filename):
    """Remove or replace characters that are invalid in filenames."""
    invalid_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(invalid_chars, '', filename)
    sanitized = sanitized.strip('. ')
    sanitized = re.sub(r'\s+', ' ', sanitized)
    return sanitized


def translate_filename(filename, client, source_language, target_language):
    """
    Translate filename from source_language to target_language.

    Args:
        filename: Original filename (without extension)
        client: Anthropic client instance
        source_language: Source language display name (e.g. 'Chinese (Simplified)')
        target_language: Target language display name (e.g. 'English')

    Returns:
        Translated filename
    """
    if source_language == target_language:
        return filename

    print(f"Translating filename: {filename}")

    context = (
        f"You are translating an Excel filename from {source_language} to {target_language}.\n"
        "This is internal business documentation from a product development company.\n"
        "Note: SDR refers to Spectrum Dynamic Research, HC refers to the R&D department.\n\n"
        f"Translate this filename to {target_language}. Keep it concise and suitable for a filename.\n"
        "Provide only the translated filename, nothing else:\n\n"
    )

    try:
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=200,
            messages=[{"role": "user", "content": context + filename}]
        )

        if message.stop_reason == 'refusal':
            print("  Warning: Translation refused, keeping original filename")
            return filename

        if message and hasattr(message, 'content') and message.content:
            if len(message.content) > 0:
                content_block = message.content[0]
                if hasattr(content_block, 'text'):
                    translated = content_block.text.strip()
                    if translated:
                        sanitized = sanitize_filename(translated)
                        print(f"Translated filename to: {sanitized}")
                        return sanitized

        print("  Warning: Could not translate filename, keeping original")
        return filename

    except Exception as e:
        print(f"  Error translating filename: {e}")
        return filename


def generate_output_filename(input_path, client, source_language, target_language):
    """
    Generate output filename by translating and adding '_translated' suffix.

    Args:
        input_path: Path to input file
        client: Anthropic client instance
        source_language: Source language display name
        target_language: Target language display name

    Returns:
        Output file path with translated filename
    """
    directory = os.path.dirname(input_path)
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)

    translated_name = translate_filename(name, client, source_language, target_language)
    output_filename = f"{translated_name}_translated{ext}"
    output_path = os.path.join(directory, output_filename)

    return output_path


def translate_text(text, client, source_language, target_language, max_retries=2):
    """
    Translate text using Claude API with retry logic.

    Args:
        text: Text to translate
        client: Anthropic client instance
        source_language: Source language display name
        target_language: Target language display name
        max_retries: Maximum number of retry attempts

    Returns:
        Translated text
    """
    context = (
        "You are translating internal business documentation from a product development company.\n"
        "This documentation contains technical specifications, sample requirements, and testing procedures.\n"
        "Note: SDR refers to Spectrum Dynamic Research, which is the testing department.\n"
        "HC refers to the R&D department.\n\n"
        f"Please translate the following {source_language} text to {target_language}. "
        f"Provide only the {target_language} translation, nothing else:\n\n"
    )

    for attempt in range(max_retries):
        try:
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=2000,
                messages=[{"role": "user", "content": context + text}]
            )

            if message.stop_reason == 'refusal':
                print(f"  Refusal on attempt {attempt + 1}/{max_retries}")
                if attempt < max_retries - 1:
                    print("  Retrying...")
                    time.sleep(2)
                    continue
                else:
                    print("  All attempts resulted in refusal")
                    return text

            if message and hasattr(message, 'content') and message.content:
                if len(message.content) > 0:
                    content_block = message.content[0]
                    if hasattr(content_block, 'text'):
                        translated = content_block.text.strip()
                        if translated:
                            return translated

            if attempt < max_retries - 1:
                time.sleep(1)

        except Exception as e:
            print(f"  Exception (attempt {attempt + 1}): {type(e).__name__}: {e}")
            if attempt < max_retries - 1:
                time.sleep(1)

    print(f"  ERROR: Translation failed for: {text[:100]}...")
    return text


def needs_translation(text, source_language, target_language):
    """
    Check if text needs translation.

    Skips translation when source == target, or when the text contains no
    alphabetic or CJK content (e.g. pure numbers, symbols, whitespace).

    Args:
        text: Text to check
        source_language: Source language display name
        target_language: Target language display name

    Returns:
        True if translation should be attempted, False otherwise
    """
    if source_language == target_language:
        return False

    stripped = text.strip()
    if not stripped:
        return False

    # Only translate if the text contains actual language characters
    has_language_content = any(
        c.isalpha() or '\u4e00' <= c <= '\u9fff' or '\u3040' <= c <= '\u30ff'
        for c in stripped
    )
    return has_language_content


def process_cell(cell, client, source_language, target_language):
    """
    Process a single cell, translating text if needed.

    Returns:
        True if cell was translated, False otherwise
    """
    try:
        if cell.value is None:
            return False

        original_text = str(cell.value).strip()

        if not original_text:
            return False

        if needs_translation(original_text, source_language, target_language):
            translated_text = translate_text(original_text, client, source_language, target_language, max_retries=2)

            if translated_text != original_text:
                cell.value = translated_text
                return True
            else:
                print("  WARNING: Translation failed for cell - original text retained")
                return False

    except Exception as e:
        print(f"  Error processing cell: {e}")
        return False

    return False


def translate_chart_text(text, client, source_language, target_language):
    """Translate text in charts if needed."""
    if not text:
        return text

    text_str = str(text).strip()
    if not text_str:
        return text

    if needs_translation(text_str, source_language, target_language):
        translated = translate_text(text_str, client, source_language, target_language, max_retries=2)
        if translated != text_str:
            return translated

    return text


def process_chart(chart, client, source_language, target_language):
    """
    Process a single chart, translating all text elements.

    Returns:
        Number of chart elements translated
    """
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import Paragraph, ParagraphProperties, RegularTextRun

    translated_count = 0

    try:
        # Translate chart title
        if chart.title:
            original_title = None

            if hasattr(chart.title, 'tx') and chart.title.tx:
                if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                    if hasattr(chart.title.tx.rich, 'p') and chart.title.tx.rich.p:
                        text_parts = []
                        for para in chart.title.tx.rich.p:
                            if hasattr(para, 'r') and para.r:
                                for run in para.r:
                                    if hasattr(run, 't') and run.t:
                                        text_parts.append(run.t)
                        original_title = ''.join(text_parts) if text_parts else None

            if original_title:
                translated_title = translate_chart_text(original_title, client, source_language, target_language)
                if translated_title != original_title:
                    chart.title.tx.rich = RichText(
                        p=[
                            Paragraph(
                                pPr=ParagraphProperties(),
                                r=[RegularTextRun(t=translated_title)]
                            )
                        ]
                    )
                    print(f"    Translated chart title: {original_title[:50]}... -> {translated_title[:50]}...")
                    translated_count += 1

        # Translate axis titles
        for axis_name, axis in [('x_axis', getattr(chart, 'x_axis', None)),
                                 ('y_axis', getattr(chart, 'y_axis', None))]:
            if axis and hasattr(axis, 'title') and axis.title:
                original_text = None

                if hasattr(axis.title, 'tx') and axis.title.tx:
                    if hasattr(axis.title.tx, 'rich') and axis.title.tx.rich:
                        if hasattr(axis.title.tx.rich, 'p') and axis.title.tx.rich.p:
                            text_parts = []
                            for para in axis.title.tx.rich.p:
                                if hasattr(para, 'r') and para.r:
                                    for run in para.r:
                                        if hasattr(run, 't') and run.t:
                                            text_parts.append(run.t)
                            original_text = ''.join(text_parts) if text_parts else None

                if original_text:
                    translated_text = translate_chart_text(original_text, client, source_language, target_language)
                    if translated_text != original_text:
                        axis.title.tx.rich = RichText(
                            p=[
                                Paragraph(
                                    pPr=ParagraphProperties(),
                                    r=[RegularTextRun(t=translated_text)]
                                )
                            ]
                        )
                        print(f"    Translated {axis_name}: {original_text[:50]}... -> {translated_text[:50]}...")
                        translated_count += 1

    except Exception as e:
        import traceback
        print(f"    Warning: Could not fully process chart - {e}")
        print(f"    Traceback: {traceback.format_exc()}")

    return translated_count


def translate_excel(input_file, output_file, source_language, target_language, client_param=None):
    """
    Translate all cells and charts in an Excel file.

    Args:
        input_file: Path to input Excel file
        output_file: Path to save translated Excel file
        source_language: Source language display name (e.g. 'Chinese (Simplified)')
        target_language: Target language display name (e.g. 'English')
        client_param: Anthropic client instance (optional, uses global if not provided)
    """
    active_client = client_param if client_param else client

    print(f"Loading workbook: {input_file}")
    print(f"Translating: {source_language} → {target_language}")

    workbook = load_workbook(input_file)

    total_cells_translated = 0
    total_charts_translated = 0

    for sheet_num, sheet_name in enumerate(workbook.sheetnames, 1):
        print(f"\nProcessing sheet {sheet_num}/{len(workbook.sheetnames)}: {sheet_name}")
        sheet = workbook[sheet_name]

        sheet_cells_translated = 0
        sheet_charts_translated = 0

        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
            print("  Sheet is empty, skipping cells")
        else:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                       min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if process_cell(cell, active_client, source_language, target_language):
                        sheet_cells_translated += 1
                        total_cells_translated += 1

        if hasattr(sheet, '_charts') and sheet._charts:
            print(f"  Found {len(sheet._charts)} chart(s) in sheet")
            for chart_num, chart in enumerate(sheet._charts, 1):
                print(f"  Processing chart {chart_num}/{len(sheet._charts)}")
                chart_translated = process_chart(chart, active_client, source_language, target_language)
                sheet_charts_translated += chart_translated
                total_charts_translated += chart_translated

        print(f"  Sheet summary: {sheet_cells_translated} cells, {sheet_charts_translated} chart elements translated")

    print(f"\nSaving translated workbook: {output_file}")
    workbook.save(output_file)
    print(f"\nTranslation complete!")
    print(f"Total: {total_cells_translated} cells and {total_charts_translated} chart elements translated")


if __name__ == "__main__":
    print("Excel Bidirectional Translator")
    print("=" * 50)

    input_excel = select_file()

    if not input_excel:
        print("No file selected. Exiting.")
    else:
        print("\nSelect source language:")
        print("1. English")
        print("2. Chinese (Simplified)")
        src_choice = input("Enter choice (1 or 2): ").strip()
        source_language = "Chinese (Simplified)" if src_choice == '2' else "English"

        print("\nSelect target language:")
        print("1. English")
        print("2. Chinese (Simplified)")
        tgt_choice = input("Enter choice (1 or 2): ").strip()
        target_language = "Chinese (Simplified)" if tgt_choice == '2' else "English"

        output_excel = generate_output_filename(input_excel, client, source_language, target_language)

        print(f"\nInput file:  {input_excel}")
        print(f"Output file: {output_excel}")
        print()

        translate_excel(input_excel, output_excel, source_language, target_language)
