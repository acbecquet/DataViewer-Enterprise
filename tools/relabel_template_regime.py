#!/usr/bin/env python3
"""v2.2.1 one-shot: relabel per-row 'Resistance (Ω)' -> 'Puffing Regime' (Excel
column E of each 12-col sample block) and pre-fill that column's data rows with
the block's header 'Puffing Regime:' value. Preserves cell styles (value-only
assignment). Idempotent. ONLY touches the Standardized .xlsx."""
import openpyxl

PATH = "resources/templates/Standardized Test Template - December 2025.xlsx"
wb = openpyxl.load_workbook(PATH)
relabelled = prefilled = 0
for ws in wb.worksheets:
    if "sop" in ws.title.lower():
        continue
    nblocks = max(1, ws.max_column // 12)
    for b in range(nblocks):
        base = b * 12                       # 0-based block start
        hdr_row = None
        for r in range(1, 13):
            v = ws.cell(row=r, column=base + 1).value
            if v and str(v).strip().lower() == "puffs":
                hdr_row = r; break
        if hdr_row is None:
            print(f"  SKIP {ws.title!r} block {b}: no 'puffs' header row")
            continue
        res_cell = ws.cell(row=hdr_row, column=base + 5)   # col E of the block
        if not (res_cell.value and "resist" in str(res_cell.value).lower()):
            continue                        # already relabelled / non-standard
        res_cell.value = "Puffing Regime"
        relabelled += 1
        regime = ws.cell(row=2, column=base + 8).value     # header 'Puffing Regime:' value
        if regime:
            r = hdr_row + 1
            while True:
                puff = ws.cell(row=r, column=base + 1).value
                if puff is None or str(puff).strip() == "":
                    break
                cell = ws.cell(row=r, column=base + 5)
                if cell.value in (None, ""):
                    cell.value = regime
                    prefilled += 1
                r += 1
wb.save(PATH)
print(f"Relabelled {relabelled} header cell(s); pre-filled {prefilled} data cell(s).")
