#!/usr/bin/env python3
"""Drift detector for the DataViewer upload sidecar.

Extracts the VBA from a deployed `Automated Testing Template.xlsm` and diffs
each module against the canonical `.bas` / `ThisWorkbook.cls.txt` in this
folder. This is the anti-divergence poka-yoke: the repo is the source of truth,
and this tells you whether the workbook actually matches it.

    python verify_sidecar.py --file "C:\\path\\to\\Automated Testing Template.xlsm"

Exit code 0 if every module matches, 1 if any differ / are missing.

Must be run with the MIP-allowlisted Python so the encrypted .xlsm reads as
plaintext. Requires `oletools` (olevba).
"""
import argparse
import difflib
import os
import sys

# deployed VBA module name -> repo file. Module1 is the pre-rename alias of
# TestingTools, so accept either name for that slot.
MODULE_MAP = {
    "DataViewerUpload": "DataViewerUpload.bas",
    "TestingTools": "TestingTools.bas",
    "Module1": "TestingTools.bas",
    "SampleNav": "SampleNav.bas",
    "ThisWorkbook": "ThisWorkbook.cls.txt",
}

# Single-sourced from the builder when it is importable (same folder); fall
# back to the shipped v1.2 values so verify stays runnable standalone.
try:
    from build_clean_template import CANON, TS_BANNER_FIRST, TS_CHECK_COL
except Exception:
    CANON = ["Lifetime Test", "User Test Simulation", "Long Puff Lifetime Test",
             "Rapid Puff Lifetime Test", "Intense Test", "Big Headspace Serial Test",
             "Negative Pressure Test", "Temperature Cycling Test #2",
             "Viscosity Compatibility", "Various Oil Compatibility",
             "Custom Test Template", "Temperature Cycling Test #1"]
    TS_BANNER_FIRST, TS_CHECK_COL = 18, 2


def assert_not_mip_ciphertext(path):
    """Duplicated from build_clean_template so verify stays standalone: the
    deployed .xlsm is MIP-encrypted at rest; only the allowlisted Python reads
    plaintext. Fail loudly instead of diffing ciphertext (audit M-k)."""
    with open(path, "rb") as f:
        head = f.read(16)
    if head.startswith(b"%TSD-Header-"):
        raise SystemExit("FATAL: %s reads as MIP ciphertext under this interpreter.\n"
                         "Run with the allowlisted Python: "
                         "C:/Users/S1134987/AppData/Local/Programs/Python/Python313/python.exe" % path)
    if not head.startswith(b"PK\x03\x04"):
        raise SystemExit("FATAL: %s does not look like an OOXML (.xlsm/.xlsx) package." % path)


def normalize(code: str) -> str:
    """Compare code from the first `Option Explicit` onward, ignoring the
    Attribute header, an explanatory comment header, trailing whitespace, and
    trailing blank lines. This makes the Module1->TestingTools rename and the
    repo's pasted ThisWorkbook header invisible to the diff."""
    lines = code.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    lines = [l for l in lines if not l.startswith("Attribute ")]
    start = next((i for i, l in enumerate(lines)
                  if l.strip() == "Option Explicit"), 0)
    lines = [l.rstrip() for l in lines[start:]]
    while lines and lines[-1] == "":
        lines.pop()
    return "\n".join(lines)


def _norm_xml(s):
    import re as _re
    return _re.sub(r"\s+", " ", s.replace("\r\n", "\n")).strip()


def check_ribbon_and_addin(xlsm_path, repo_dir):
    """Returns (any_diff, lines). Compares the workbook's customUI14.xml to the
    repo copy (normalized) and asserts no web-extension (add-in) parts remain."""
    import zipfile
    lines, any_diff = [], False
    try:
        with zipfile.ZipFile(xlsm_path) as z:
            names = z.namelist()
            wb_ui = z.read("customUI/customUI14.xml").decode("utf-8") \
                if "customUI/customUI14.xml" in names else ""
            webext = [n for n in names if n.startswith("xl/webextensions/")]
    except Exception as e:  # pragma: no cover
        return True, ["[!] could not read workbook zip: %r" % e]
    repo_ui = open(os.path.join(repo_dir, "customUI14.xml"), encoding="utf-8").read()
    if not wb_ui:
        any_diff = True
        lines.append("[!] customUI14.xml: MISSING from workbook")
    elif _norm_xml(wb_ui) == _norm_xml(repo_ui):
        lines.append("[OK]      customUI14.xml == repo")
    else:
        any_diff = True
        lines.append("[DIFFERS] customUI14.xml != repo")
    if webext:
        any_diff = True
        lines.append("[!] web add-in parts present: %s" % ", ".join(webext))
    else:
        lines.append("[OK]      no web-extension/add-in parts")
    return any_diff, lines


def check_workbook_structure(xlsm_path):
    """openpyxl: Test Selection present, _Settings very hidden, no 'DataViewer Upload',
    named ranges resolve to the right sheets, and the 13 selection names are in order."""
    lines, any_diff = [], False
    EXPECTED = ["Custom Test Template", "Lifetime Test", "Long Puff Lifetime Test",
                "Rapid Puff Lifetime Test", "Intense Test", "User Test Simulation",
                "Big Headspace Serial Test", "Viscosity Compatibility",
                "Various Oil Compatibility", "Temperature Cycling Test #1",
                "Temperature Cycling Test #2", "Negative Pressure Test", "Test SOP's"]
    DEST = {"DV_TestSelection": "Test Selection", "DV_FileName": "_Settings",
            "DV_SynologyPath": "_Settings", "DV_LocalPath": "_Settings",
            "DV_DataViewerExe": "_Settings", "DV_Status": "_Settings", "DV_Log": "_Settings",
            "DV_OrigFileName": "_Settings", "DV_LastUpload": "_Settings!$B$8"}
    try:
        import openpyxl
    except Exception as e:
        return False, ["[i] openpyxl unavailable; structure check skipped (%r)" % e]
    try:
        wb = openpyxl.load_workbook(xlsm_path, read_only=False, keep_vba=True)
    except Exception as e:
        return True, ["[!] could not open workbook: %r" % e]

    def ok(cond, msg):
        nonlocal any_diff
        lines.append(("[OK]      " if cond else "[DIFFERS] ") + msg)
        if not cond:
            any_diff = True

    sn = wb.sheetnames
    ok("Test Selection" in sn, "'Test Selection' sheet present")
    ok("DataViewer Upload" not in sn, "old 'DataViewer Upload' sheet removed")
    ok("_Settings" in sn, "_Settings sheet present")
    try:
        ok(wb["_Settings"].sheet_state == "veryHidden", "_Settings is very hidden")
    except Exception:
        ok(False, "_Settings readable")

    # v1.2 presence (audit M-f): every canonical, snapshot, and system sheet
    # must exist -- a build that silently drops one must not verify clean.
    required = list(CANON) + ["_Template_%02d" % i for i in range(12)] + \
        ["_Template_Master", "Test SOP's", "Test Selection", "_Settings"]
    missing = [s for s in required if s not in sn]
    ok(not missing, "all %d required sheets present" % len(required) +
       ("" if not missing else " -- missing: %s" % ", ".join(missing)))

    # v1.2 visibility: the three working sheets visible, the other canonicals
    # hidden (un-hidden by ticking their checkbox), plumbing veryHidden.
    VISIBLE = {"Test Selection", "Test SOP's", "Lifetime Test"}
    wrong = []
    for s in required:
        if s not in sn:
            continue
        want = ("veryHidden" if s.startswith("_Template_") or s == "_Settings"
                else "visible" if s in VISIBLE else "hidden")
        got = wb[s].sheet_state
        if got != want:
            wrong.append("%s=%s (want %s)" % (s, got, want))
    ok(not wrong, "sheet visibility (3 visible / canonicals hidden / _* veryHidden)" +
       ("" if not wrong else " -- " + "; ".join(wrong)))

    # defined-name destinations (version-tolerant access)
    dn = wb.defined_names
    def dest_of(name):
        try:
            return dn[name].value
        except Exception:
            try:
                return dict(dn.items())[name].value
            except Exception:
                return None
    for nm, sheet in DEST.items():
        d = dest_of(nm)
        # strip sheet-name quoting so "_Settings!$B$8" matches both
        # '_Settings'!$B$8 and _Settings!$B$8
        ok(d is not None and sheet in d.replace("'", ""),
           "%s -> %s (%r)" % (nm, sheet, d))

    if "Test Selection" in sn:
        ws = wb["Test Selection"]
        # Read the labels at the geometry the build writes them to -- single-sourced
        # via build_clean_template's TS_* constants so this check tracks the table's
        # margin offsets instead of a hardcoded A3:B15 it would silently outlive.
        fr, lc, start = 4, 3, "$B$4"        # fallback: the shipped B4:C16 layout
        try:
            import build_clean_template as _B
            fr, lc = _B.TS_FIRST_ROW, _B.TS_LABEL_COL
            start = "$%s$%d" % (_B._col_letter(_B.TS_CHECK_COL), _B.TS_FIRST_ROW)
        except Exception:
            pass
        got = [ws.cell(fr + i, lc).value for i in range(len(EXPECTED))]
        ok(got == EXPECTED,
           "Test Selection labels in canonical order (col %d, rows %d-%d)"
           % (lc, fr, fr + len(EXPECTED) - 1))
        d = dest_of("DV_TestSelection")
        ok(d is not None and start in d.replace(" ", ""),
           "DV_TestSelection anchored at %s (%r)" % (start, d))
        # M-f: the derived check above moves WITH the imported geometry, so it
        # can drift in lockstep with a broken build. Pin the exact shipped ref
        # verbatim as well (string equality, not substring -- $B$40 must fail).
        LIT = "'Test Selection'!$B$4:$C$16"
        ok(d == LIT, "DV_TestSelection RefersTo == %s exactly (%r)" % (LIT, d))
        # v1.2 banner (P5): the 'never email' line must be on-sheet -- the only
        # guidance channel that works with macros disabled. The banner rows are
        # merged; openpyxl surfaces the value in the top-left cell, which is
        # (row, TS_CHECK_COL).
        bv = ws.cell(TS_BANNER_FIRST + 1, TS_CHECK_COL).value
        ok(isinstance(bv, str) and "never email" in bv.lower(),
           "banner 'never email' line at row %d col %d (%r)"
           % (TS_BANNER_FIRST + 1, TS_CHECK_COL, bv))
    return any_diff, lines


def check_clog_formatting(xlsm_path):
    """openpyxl: the v1.1 Feature 3 auto-Clog formula + conditional format are
    baked onto a known block-layout sheet ('Custom Test Template', 4 blocks ->
    Clog col G in block 1). Returns (any_diff, lines)."""
    lines, any_diff = [], False

    def ok(cond, msg):
        nonlocal any_diff
        lines.append(("[OK]      " if cond else "[DIFFERS] ") + msg)
        if not cond:
            any_diff = True

    try:
        from openpyxl import load_workbook
    except Exception as e:
        return False, ["[i] openpyxl unavailable; Clog check skipped (%r)" % e]
    try:
        wb = load_workbook(xlsm_path, data_only=False)
        ws = wb["Custom Test Template"]      # 4 blocks; Clog col G (block 1)
    except Exception as e:
        return True, ["[!] could not read Clog sheet: %r" % e]
    # v1.2: expect the ISNUMBER form (audit M-d -- text like 'n/a' must yield
    # EMPTY, not Heavy Clog). Single-sourced from the builder when importable.
    try:
        from build_clean_template import clog_formula as _clog
        want = _clog("D", 5)
    except Exception:
        want = ('=IF(NOT(ISNUMBER(D5)),"",IF(D5>=15,"Heavy Clog",'
                'IF(D5>5,"Light Clog","")))')
    f = ws["G5"].value
    ok(isinstance(f, str) and f.replace(" ", "") == want.replace(" ", ""),
       "Clog formula at G5 is the ISNUMBER form (Custom Test Template) (%r)" % f)
    ok(len(list(ws.conditional_formatting)) >= 1, "Clog conditional formatting present (Custom Test Template)")
    return any_diff, lines


def check_ribbon_icons(xlsm_path):
    """Ribbon customUI images must be sanely sized -- a 5120x5120 imgPlus made
    Excel raise 'image too large'. Asserts every customUI PNG is <= 1024 px in
    each dimension. Returns (any_diff, lines)."""
    import struct
    import zipfile as _zip
    lines, any_diff = [], False

    def ok(cond, msg):
        nonlocal any_diff
        lines.append(("[OK]      " if cond else "[DIFFERS] ") + msg)
        if not cond:
            any_diff = True

    LIMIT = 1024
    try:
        z = _zip.ZipFile(xlsm_path)
    except Exception as e:
        return True, ["[!] could not open package for ribbon-icon check: %r" % e]
    imgs = [n for n in z.namelist()
            if n.startswith("customUI/images/") and n.lower().endswith(".png")]
    if not imgs:
        return False, ["[i] no customUI ribbon images found"]
    for n in sorted(imgs):
        b = z.read(n)
        if b[:8] == b"\x89PNG\r\n\x1a\n" and b[12:16] == b"IHDR":
            w, h = struct.unpack(">II", b[16:24])
            ok(w <= LIMIT and h <= LIMIT,
               "ribbon icon %s is %dx%d (<= %d px)" % (n.split("/")[-1], w, h, LIMIT))
    return any_diff, lines


def check_package_parts(xlsm_path, require_signature=False):
    """Zip-level part presence. featurePropertyBag carries the native (real)
    checkboxes -- if those parts are lost the boxes silently degrade to
    TRUE/FALSE text (audit M-f). The VBA signature part is only enforced under
    --require-signature (the runbook's final post-sign gate); the pre-sign
    build check runs without it. Returns (any_diff, lines)."""
    import zipfile as _zip
    lines, any_diff = [], False

    def ok(cond, msg):
        nonlocal any_diff
        lines.append(("[OK]      " if cond else "[DIFFERS] ") + msg)
        if not cond:
            any_diff = True

    try:
        with _zip.ZipFile(xlsm_path) as z:
            names = z.namelist()
    except Exception as e:
        return True, ["[!] could not open package for part checks: %r" % e]
    ok(any(n.startswith("xl/featurePropertyBag") for n in names),
       "featurePropertyBag parts present (native checkbox carrier)")
    signed = any(n.startswith("xl/vbaProjectSignature") for n in names)
    if require_signature:
        ok(signed, "VBA project signature present (--require-signature)")
    else:
        lines.append("[i] VBA signature not required (signature %s); pass "
                     "--require-signature to enforce"
                     % ("present" if signed else "absent"))
    return any_diff, lines


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--file", required=True, help="path to the .xlsm to check")
    ap.add_argument("--repo", default=os.path.dirname(os.path.abspath(__file__)),
                    help="folder with the canonical files (default: this folder)")
    ap.add_argument("--require-signature", action="store_true",
                    help="fail unless the VBA project carries a digital "
                         "signature (xl/vbaProjectSignature* part)")
    args = ap.parse_args()

    if not os.path.isfile(args.file):
        print("ERROR: file not found:", args.file)
        return 1
    assert_not_mip_ciphertext(args.file)
    try:
        from oletools.olevba import VBA_Parser
    except Exception as e:  # pragma: no cover
        print("ERROR: oletools (olevba) not available:", e)
        return 1

    # Extract deployed modules.
    deployed = {}
    vp = VBA_Parser(args.file)
    try:
        for _fn, _stream, vba_name, code in vp.extract_macros():
            deployed[os.path.splitext(vba_name)[0]] = code
    finally:
        vp.close()

    print("Deployed modules:", ", ".join(sorted(deployed)) or "(none)")
    print("-" * 60)

    any_diff = False
    checked_files = set()
    for dep_name, code in sorted(deployed.items()):
        repo_file = MODULE_MAP.get(dep_name)
        if repo_file is None:
            # Sheet code modules (Sheet1.cls etc.) are stubs - skip unless they
            # carry real code.
            if normalize(code):
                print(f"[?] {dep_name}: in workbook but unmanaged by the repo "
                      f"(has code) - review")
                any_diff = True
            continue
        repo_path = os.path.join(args.repo, repo_file)
        checked_files.add(repo_file)
        if not os.path.isfile(repo_path):
            print(f"[!] {dep_name}: repo file {repo_file} missing")
            any_diff = True
            continue
        # VBA identifiers are case-insensitive, and the VBE auto-recases them to
        # one canonical spelling per project on import (e.g. a 'Dim names' local
        # forces every '.Names' property reference to '.names'). Compare
        # case-insensitively so cosmetic recasing isn't reported as drift; a real
        # change still differs.
        dep_n = normalize(code).lower()
        repo_n = normalize(open(repo_path, encoding="utf-8").read()).lower()
        if dep_n == repo_n:
            print(f"[OK] {dep_name:18} == {repo_file}")
        else:
            any_diff = True
            print(f"[DIFFERS] {dep_name:18} != {repo_file}")
            diff = difflib.unified_diff(dep_n.split("\n"), repo_n.split("\n"),
                                        "deployed", "repo", lineterm="", n=0)
            shown = [l for l in diff if l.startswith(("+", "-"))
                     and not l.startswith(("+++", "---"))]
            for l in shown[:24]:
                print("    " + l)
            if len(shown) > 24:
                print(f"    ... ({len(shown) - 24} more changed lines)")

    # repo files with no matching deployed module
    for repo_file in sorted(set(MODULE_MAP.values()) - checked_files):
        print(f"[!] {repo_file}: no matching module found in the workbook")
        any_diff = True

    rib_diff, rib_lines = check_ribbon_and_addin(args.file, args.repo)
    print("-" * 60)
    for l in rib_lines:
        print(l)
    any_diff = any_diff or rib_diff

    st_diff, st_lines = check_workbook_structure(args.file)
    for l in st_lines:
        print(l)
    any_diff = any_diff or st_diff

    clog_diff, clog_lines = check_clog_formatting(args.file)
    for l in clog_lines:
        print(l)
    any_diff = any_diff or clog_diff

    icon_diff, icon_lines = check_ribbon_icons(args.file)
    for l in icon_lines:
        print(l)
    any_diff = any_diff or icon_diff

    part_diff, part_lines = check_package_parts(args.file, args.require_signature)
    for l in part_lines:
        print(l)
    any_diff = any_diff or part_diff

    print("-" * 60)
    print("RESULT:", "DRIFT DETECTED" if any_diff else "all modules match")
    return 1 if any_diff else 0


if __name__ == "__main__":
    sys.exit(main())
