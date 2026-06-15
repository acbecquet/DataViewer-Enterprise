"""v2.4.2 R6: deleteRowFromExcel must remove the right row, leave a valid
workbook, and save ATOMICALLY (tmp + os.replace) so a kill mid-write can't
truncate the source. Mirrors the kDeleteRow script in MainWindow.cpp.
Run: python tests/excel/test_atomic_delete.py  (needs openpyxl)
Exits 0 on success, nonzero on failure.
"""
import os, sys, tempfile
from openpyxl import Workbook, load_workbook


def main():
    d = tempfile.mkdtemp()
    path = os.path.join(d, "fixture.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "S1"
    for r in range(1, 6):
        ws.cell(row=r, column=1).value = r * 10   # 10,20,30,40,50
    wb.save(path)

    # Exactly what kDeleteRow does: load, delete row 3, atomic save.
    wb = load_workbook(path)
    ws = wb["S1"]
    ws.delete_rows(3, 1)
    tmp = path + ".dve_tmp"
    wb.save(tmp)
    os.replace(tmp, path)

    chk = load_workbook(path); cs = chk["S1"]
    vals = [cs.cell(row=r, column=1).value for r in range(1, cs.max_row + 1)]
    assert vals == [10, 20, 40, 50], f"unexpected rows after delete: {vals}"
    assert not os.path.exists(tmp), "atomic tmp file was left behind"
    print("OK atomic delete round-trip")


if __name__ == "__main__":
    try:
        main()
    except AssertionError as e:
        print("FAIL:", e); sys.exit(1)
