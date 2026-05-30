#!/usr/bin/env python3
"""Generate test .xlsx fixtures for DataViewer Enterprise test suite.

Each fixture is minimal (2 samples, 5 data rows) with realistic TPM data.
Run: python generate_fixtures.py
Output: tests/data/*.xlsx + tests/data/test_image.png
"""
import os, sys

try:
    from openpyxl import Workbook
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl")
    sys.exit(1)

DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(DATA_DIR, exist_ok=True)

# Realistic TPM test data: 5 intervals, puffs in increments of 10
PUFFS    = [10, 20, 30, 40, 50]
BEFORE_W = [25.1000, 25.0650, 25.0320, 24.9960, 24.9630]
AFTER_W  = [25.0650, 25.0320, 24.9960, 24.9630, 24.9280]
# TPM = (before - after) * 1000 / interval = 3.5 mg/puff each
DRAW_P   = [0.45, 0.42, 0.44, 0.43, 0.41]

PUFFS2   = [10, 20, 30, 40, 50]
BEFORE2  = [26.2000, 26.1580, 26.1200, 26.0790, 26.0410]
AFTER2   = [26.1580, 26.1200, 26.0790, 26.0410, 26.0010]
DRAW_P2  = [0.50, 0.48, 0.51, 0.49, 0.47]


def write_format_e_headers(ws, off, sample_id, date="2026-03-16", heating="CCELL3.0"):
    """Write Format E (Dec 2025) metadata and headers starting at column offset."""
    o = off  # 1-based column offset
    # Row 1: metadata
    ws.cell(row=1, column=o+1, value="Lifetime Test")
    ws.cell(row=1, column=o+3, value="Date:")
    ws.cell(row=1, column=o+4, value=date)
    ws.cell(row=1, column=o+5, value="Sample ID:")
    ws.cell(row=1, column=o+6, value=sample_id)
    ws.cell(row=1, column=o+7, value="Heating Technology:")
    ws.cell(row=1, column=o+8, value=heating)
    ws.cell(row=1, column=o+10, value="Did this burn?")
    # Row 2
    ws.cell(row=2, column=o+1, value="Media:")
    ws.cell(row=2, column=o+2, value="D9")
    ws.cell(row=2, column=o+3, value="Resistance (\u03A9):")
    ws.cell(row=2, column=o+4, value=1.1)
    ws.cell(row=2, column=o+5, value="Power:")
    ws.cell(row=2, column=o+6, value=4.17)
    ws.cell(row=2, column=o+7, value="Puffing Regime:")
    ws.cell(row=2, column=o+8, value="60mL/3s/30s")
    ws.cell(row=2, column=o+9, value="Usage Efficiency")
    ws.cell(row=2, column=o+10, value="Did this clog?")
    # Row 3
    ws.cell(row=3, column=o+1, value="Viscosity:")
    ws.cell(row=3, column=o+2, value=500000)
    ws.cell(row=3, column=o+3, value="Tester:")
    ws.cell(row=3, column=o+4, value="TestUser")
    ws.cell(row=3, column=o+5, value="Voltage:")
    ws.cell(row=3, column=o+6, value=2.8)
    ws.cell(row=3, column=o+7, value="Initial Oil Mass:")
    ws.cell(row=3, column=o+8, value=1.0)
    ws.cell(row=3, column=o+10, value="Did this leak?")
    # Row 4: headers
    hdrs = ["puffs", "Before weight (g)", "After weight (g)", "Draw Pressure (kpa)",
            "Resistance (\u03A9)", "Smell", "Clog", "Notes",
            "TPM (mg/puff)", "TPM Power Density (mg/(W*s))",
            "Variation in TPM (%)", "Oil Consumed (Cumulative, g)"]
    for i, h in enumerate(hdrs):
        ws.cell(row=4, column=o+1+i, value=h)


def write_data_rows(ws, off, puffs, before, after, draw, start_row=5):
    for i in range(len(puffs)):
        r = start_row + i
        o = off
        ws.cell(row=r, column=o+1, value=puffs[i])
        ws.cell(row=r, column=o+2, value=before[i])
        ws.cell(row=r, column=o+3, value=after[i])
        ws.cell(row=r, column=o+4, value=draw[i])


# ──────────────────────────────────────────────────────────────────────
# Format E — Dec 2025 (current, native)
# ──────────────────────────────────────────────────────────────────────
def gen_format_e():
    wb = Workbook()
    # Sheet 1: Lifetime Test
    ws = wb.active
    ws.title = "Lifetime Test"
    write_format_e_headers(ws, 0, "Lifetime-1")
    write_data_rows(ws, 0, PUFFS, BEFORE_W, AFTER_W, DRAW_P)
    write_format_e_headers(ws, 12, "Lifetime-2")
    write_data_rows(ws, 12, PUFFS2, BEFORE2, AFTER2, DRAW_P2)

    # Sheet 2: Long Puff Lifetime Test (triggers "new" template detection)
    ws2 = wb.create_sheet("Long Puff Lifetime Test")
    write_format_e_headers(ws2, 0, "LongPuff-1")
    write_data_rows(ws2, 0, [2, 4, 6, 8, 10], BEFORE_W, AFTER_W, DRAW_P)

    wb.save(os.path.join(DATA_DIR, "format_e.xlsx"))
    print("  format_e.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Format D — Jan 2025 standard (12-col sheet + 11-col sheet)
# ──────────────────────────────────────────────────────────────────────
def gen_format_d():
    wb = Workbook()
    ws = wb.active
    ws.title = "Lifetime Test"
    # Same as Format E but "Resistance (Ohms):" and no Heating Technology
    o = 0
    ws.cell(row=1, column=1, value="Lifetime Test")
    ws.cell(row=1, column=3, value="Date:")
    ws.cell(row=1, column=4, value="2025-06-24")
    ws.cell(row=1, column=5, value="Sample ID:")
    ws.cell(row=1, column=6, value="Life-1")
    ws.cell(row=1, column=10, value="Did this burn?")
    ws.cell(row=1, column=12, value="Average TPM and Standard deviation")
    ws.cell(row=2, column=1, value="Media:")
    ws.cell(row=2, column=2, value="D8")
    ws.cell(row=2, column=3, value="Resistance (Ohms):")
    ws.cell(row=2, column=4, value=1.35)
    ws.cell(row=2, column=5, value="Power:")
    ws.cell(row=2, column=7, value="Puffing Regime:")
    ws.cell(row=2, column=8, value="60mL/3s/30s")
    ws.cell(row=2, column=9, value="Usage Efficiency")
    ws.cell(row=2, column=10, value="Did this clog?")
    ws.cell(row=3, column=1, value="Viscosity:")
    ws.cell(row=3, column=2, value=500000)
    ws.cell(row=3, column=3, value="Tester:")
    ws.cell(row=3, column=4, value="Nellie")
    ws.cell(row=3, column=5, value="Voltage:")
    ws.cell(row=3, column=6, value=3.4)
    ws.cell(row=3, column=7, value="Initial Oil Mass:")
    ws.cell(row=3, column=8, value=2)
    ws.cell(row=3, column=10, value="Did this leak?")
    hdrs12 = ["puffs", "Before weight/g", "After weight/g", "Draw Pressure (kpa)",
              "Resistance", "Smell", "Clog", "Notes",
              "TPM (mg/puff)", "TPM Power Density (mg/puff/W)",
              "TPM Consistency", "Rolling Average TPM"]
    for i, h in enumerate(hdrs12):
        ws.cell(row=4, column=1+i, value=h)
    write_data_rows(ws, 0, PUFFS, BEFORE_W, AFTER_W, DRAW_P)

    # Second sample
    o2 = 12
    ws.cell(row=1, column=o2+1, value="Lifetime Test")
    ws.cell(row=1, column=o2+3, value="Date:")
    ws.cell(row=1, column=o2+4, value="2025-06-24")
    ws.cell(row=1, column=o2+5, value="Sample ID:")
    ws.cell(row=1, column=o2+6, value="Life-2")
    ws.cell(row=2, column=o2+1, value="Media:")
    ws.cell(row=2, column=o2+2, value="D8")
    ws.cell(row=2, column=o2+3, value="Resistance (Ohms):")
    ws.cell(row=2, column=o2+4, value=1.35)
    ws.cell(row=2, column=o2+7, value="Puffing Regime:")
    ws.cell(row=2, column=o2+8, value="60mL/3s/30s")
    ws.cell(row=3, column=o2+1, value="Viscosity:")
    ws.cell(row=3, column=o2+2, value=500000)
    ws.cell(row=3, column=o2+3, value="Tester:")
    ws.cell(row=3, column=o2+4, value="Nellie")
    ws.cell(row=3, column=o2+5, value="Voltage:")
    ws.cell(row=3, column=o2+6, value=3.4)
    ws.cell(row=3, column=o2+7, value="Initial Oil Mass:")
    ws.cell(row=3, column=o2+8, value=2)
    for i, h in enumerate(hdrs12):
        ws.cell(row=4, column=o2+1+i, value=h)
    write_data_rows(ws, o2, PUFFS2, BEFORE2, AFTER2, DRAW_P2)

    # 11-col Long Puff variant (no Rolling Average TPM column)
    ws2 = wb.create_sheet("Long Puff Test")
    ws2.cell(row=1, column=1, value="Long Puff Test")
    ws2.cell(row=1, column=3, value="Date:")
    ws2.cell(row=1, column=5, value="Sample ID:")
    ws2.cell(row=1, column=6, value="LP-1")
    ws2.cell(row=1, column=10, value="Did this burn?")
    ws2.cell(row=2, column=1, value="Media:")
    ws2.cell(row=2, column=2, value="D8")
    ws2.cell(row=2, column=3, value="Resistance (Ohms):")
    ws2.cell(row=2, column=4, value=1.25)
    ws2.cell(row=2, column=5, value="Power:")
    ws2.cell(row=2, column=7, value="Initial Oil Mass:")
    ws2.cell(row=2, column=8, value=2)
    ws2.cell(row=2, column=9, value="Usage Efficiency")
    ws2.cell(row=2, column=10, value="Did this clog?")
    ws2.cell(row=3, column=1, value="Viscosity:")
    ws2.cell(row=3, column=2, value="68kcp")
    ws2.cell(row=3, column=3, value="Tester:")
    ws2.cell(row=3, column=5, value="Voltage:")
    ws2.cell(row=3, column=6, value=3.4)
    ws2.cell(row=3, column=10, value="Did this leak?")
    hdrs11 = ["puffs", "Before weight/g", "After weight/g", "Draw Pressure (kpa)",
              "Resistance", "Smell", "Clog", "Notes",
              "TPM (mg/puff)", "TPM Power Density (mg/puff/W)", "TPM Consistency"]
    for i, h in enumerate(hdrs11):
        ws2.cell(row=4, column=1+i, value=h)
    write_data_rows(ws2, 0, [2, 4, 6, 8, 10], BEFORE_W, AFTER_W, DRAW_P)

    wb.save(os.path.join(DATA_DIR, "format_d.xlsx"))
    print("  format_d.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Format C — Gembox/T58G (metadata remap, Project+Sample)
# ──────────────────────────────────────────────────────────────────────
def gen_format_c():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for s, (sid, proj) in enumerate([("Intense-1", "Gembox"), ("Intense-2", "Gembox")]):
        o = s * 12
        ws.cell(row=1, column=o+1, value="TPM ")
        ws.cell(row=1, column=o+2, value="Date:")
        ws.cell(row=1, column=o+3, value="2025-01-30")
        ws.cell(row=1, column=o+4, value="Tester:")
        ws.cell(row=1, column=o+5, value="Nellie")
        ws.cell(row=1, column=o+6, value="Project:")
        ws.cell(row=1, column=o+7, value=proj)
        ws.cell(row=1, column=o+8, value="Sample:")
        ws.cell(row=1, column=o+9, value=sid)
        ws.cell(row=1, column=o+10, value="Did this burn?")
        ws.cell(row=2, column=o+1, value="Media:")
        ws.cell(row=2, column=o+2, value="D8")
        ws.cell(row=2, column=o+3, value="Ri (Ohms)")
        ws.cell(row=2, column=o+4, value=2.4)
        ws.cell(row=2, column=o+5, value="Power:")
        ws.cell(row=2, column=o+6, value=4.27)
        ws.cell(row=2, column=o+7, value="Puffing Regime:")
        ws.cell(row=2, column=o+8, value="200mL/3s/30s")
        ws.cell(row=2, column=o+9, value="Usage Efficiency")
        ws.cell(row=2, column=o+10, value="Did this clog?")
        ws.cell(row=3, column=o+1, value="Viscosity:")
        ws.cell(row=3, column=o+2, value="800kcp")
        ws.cell(row=3, column=o+3, value="Rf (Ohms)")
        ws.cell(row=3, column=o+5, value="Voltage:")
        ws.cell(row=3, column=o+6, value=3.2)
        ws.cell(row=3, column=o+7, value="Initial Oil Mass:")
        ws.cell(row=3, column=o+8, value=2)
        ws.cell(row=3, column=o+10, value="Did this leak?")
        hdrs = ["puffs", "Before weight/g", "After weight/g", "Draw Pressure (kpa)",
                "Resistance", "Smell", "Clog", "Notes",
                "TPM (mg/puff)", "TPM Power Density (mg/puff/W)",
                "TPM Consistency", "Rolling Average TPM"]
        for i, h in enumerate(hdrs):
            ws.cell(row=4, column=o+1+i, value=h)
        data = (PUFFS, BEFORE_W, AFTER_W, DRAW_P) if s == 0 else (PUFFS2, BEFORE2, AFTER2, DRAW_P2)
        write_data_rows(ws, o, *data)
    wb.save(os.path.join(DATA_DIR, "format_c.xlsx"))
    print("  format_c.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Format B — M1 Extended (row 0 empty, PV1-5, no Clog, 12 cols)
# ──────────────────────────────────────────────────────────────────────
def gen_format_b():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for s, sid in enumerate(["CPS1910-1", "CPS1910-2"]):
        o = s * 12
        # Row 1 empty
        ws.cell(row=2, column=o+1, value="Cart #")
        ws.cell(row=2, column=o+2, value=sid)
        ws.cell(row=2, column=o+3, value="Ri (Ohms)")
        ws.cell(row=2, column=o+4, value=1.53)
        ws.cell(row=2, column=o+5, value="Power")
        ws.cell(row=2, column=o+7, value="Viscosity")
        ws.cell(row=2, column=o+8, value=250000)
        ws.cell(row=2, column=o+9, value="Temperature")
        ws.cell(row=2, column=o+11, value="Average TPM")
        ws.cell(row=3, column=o+1, value="Media")
        ws.cell(row=3, column=o+2, value="Distillate")
        ws.cell(row=3, column=o+3, value="Rf (Ohms)")
        ws.cell(row=3, column=o+5, value="Puff Regime")
        ws.cell(row=3, column=o+6, value="60mL/3s/30s")
        ws.cell(row=3, column=o+7, value="Voltage")
        ws.cell(row=3, column=o+8, value=3.7)
        ws.cell(row=3, column=o+11, value="Standard Deviation (TPM)")
        hdrs = ["puffs", "Before weight/g", "After weight/g", "PV1", "PV2",
                "PV3", "PV4", "PV5", "Resistance", "Smell (0-4)", "Notes", "TPM (mg/puff)"]
        for i, h in enumerate(hdrs):
            ws.cell(row=4, column=o+1+i, value=h)
        data_p = PUFFS if s == 0 else PUFFS2
        data_b = BEFORE_W if s == 0 else BEFORE2
        data_a = AFTER_W if s == 0 else AFTER2
        data_d = DRAW_P if s == 0 else DRAW_P2
        for i in range(5):
            r = 5 + i
            ws.cell(row=r, column=o+1, value=data_p[i])
            ws.cell(row=r, column=o+2, value=data_b[i])
            ws.cell(row=r, column=o+3, value=data_a[i])
            ws.cell(row=r, column=o+4, value=data_d[i])  # PV1
            # TPM at col 12
            interval = data_p[i] if i == 0 else data_p[i] - data_p[i-1]
            tpm = (data_b[i] - data_a[i]) * 1000 / interval
            ws.cell(row=r, column=o+12, value=tpm)
    wb.save(os.path.join(DATA_DIR, "format_b.xlsx"))
    print("  format_b.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Format A — Comparison Test (13 cols, PV1-5 + Clog?)
# ──────────────────────────────────────────────────────────────────────
def gen_format_a():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for s, sid in enumerate(["T51P-C39-1", "T51P-C39-2"]):
        o = s * 13
        ws.cell(row=1, column=o+1, value="Date:")
        ws.cell(row=1, column=o+3, value="Coil Material")
        ws.cell(row=1, column=o+5, value="Thermal Conductivity")
        ws.cell(row=1, column=o+9, value="Total Oil Puffed (g)")
        ws.cell(row=1, column=o+11, value="Did this burn?")
        ws.cell(row=1, column=o+13, value="Average TPM and Standard deviation")
        ws.cell(row=2, column=o+1, value="Cart #")
        ws.cell(row=2, column=o+2, value=sid)
        ws.cell(row=2, column=o+3, value="Ri (Ohms)")
        ws.cell(row=2, column=o+4, value=1.2)
        ws.cell(row=2, column=o+5, value="Power")
        ws.cell(row=2, column=o+7, value="Viscosity")
        ws.cell(row=2, column=o+8, value=500000)
        ws.cell(row=2, column=o+9, value="Did this clog?")
        ws.cell(row=2, column=o+11, value="Initial Oil Mass")
        ws.cell(row=2, column=o+12, value=1.5)
        ws.cell(row=3, column=o+1, value="Media")
        ws.cell(row=3, column=o+2, value="D9")
        ws.cell(row=3, column=o+3, value="Rf (Ohms)")
        ws.cell(row=3, column=o+5, value="Puff Regime")
        ws.cell(row=3, column=o+6, value="60mL/3s/30s")
        ws.cell(row=3, column=o+7, value="Voltage")
        ws.cell(row=3, column=o+8, value=3.4)
        ws.cell(row=3, column=o+9, value="Did this leak?")
        ws.cell(row=3, column=o+11, value="Usage Efficiency")
        hdrs = ["puffs", "Before weight/g", "After weight/g", "PV1", "PV2",
                "PV3", "PV4", "PV5", "Resistance", "Smell (0-4)", "Clog?",
                "Notes", "TPM (mg/puff)"]
        for i, h in enumerate(hdrs):
            ws.cell(row=4, column=o+1+i, value=h)
        data_p = PUFFS if s == 0 else PUFFS2
        data_b = BEFORE_W if s == 0 else BEFORE2
        data_a = AFTER_W if s == 0 else AFTER2
        data_d = DRAW_P if s == 0 else DRAW_P2
        for i in range(5):
            r = 5 + i
            ws.cell(row=r, column=o+1, value=data_p[i])
            ws.cell(row=r, column=o+2, value=data_b[i])
            ws.cell(row=r, column=o+3, value=data_a[i])
            ws.cell(row=r, column=o+4, value=data_d[i])  # PV1
            interval = data_p[i] if i == 0 else data_p[i] - data_p[i-1]
            tpm = (data_b[i] - data_a[i]) * 1000 / interval
            ws.cell(row=r, column=o+13, value=tpm)
    wb.save(os.path.join(DATA_DIR, "format_a.xlsx"))
    print("  format_a.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Empty file
# ──────────────────────────────────────────────────────────────────────
def gen_empty():
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(os.path.join(DATA_DIR, "empty.xlsx"))
    print("  empty.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Multi-sheet file (Format E with multiple test types)
# ──────────────────────────────────────────────────────────────────────
def gen_multi_sheet():
    wb = Workbook()
    for idx, sname in enumerate(["Lifetime Test", "Long Puff Lifetime Test", "Intense Test"]):
        ws = wb.active if idx == 0 else wb.create_sheet(sname)
        if idx == 0:
            ws.title = sname
        write_format_e_headers(ws, 0, f"{sname.split()[0]}-1")
        write_data_rows(ws, 0, PUFFS, BEFORE_W, AFTER_W, DRAW_P)
    wb.save(os.path.join(DATA_DIR, "multi_sheet.xlsx"))
    print("  multi_sheet.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Format E + per-row Puffing Regime (v2.2.1 new-template variant)
# ──────────────────────────────────────────────────────────────────────
def gen_format_e_regime():
    """New template (v2.2.1): per-row Puffing Regime column instead of Resistance."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lifetime Test"
    write_format_e_headers(ws, 0, "Regime-1")
    ws.cell(row=4, column=5, value="Puffing Regime")   # col-E header -> new label
    write_data_rows(ws, 0, PUFFS, BEFORE_W, AFTER_W, DRAW_P)
    regimes = ["60mL/3s/30s", "60mL/3s/30s", "60mL/3s/30s", "200mL/9s/300s", "200mL/9s/300s"]
    for i, rg in enumerate(regimes):
        ws.cell(row=5 + i, column=5, value=rg)         # per-row regimes (mid-session change)
    wb.save(os.path.join(DATA_DIR, "format_e_regime.xlsx"))
    print("  format_e_regime.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Test image (100x100 red square PNG)
# ──────────────────────────────────────────────────────────────────────
def gen_test_image():
    import struct, zlib
    # Minimal 100x100 red PNG
    w, h = 100, 100
    raw = b''
    for _ in range(h):
        raw += b'\x00'  # filter byte
        raw += b'\xff\x00\x00' * w  # RGB red
    compressed = zlib.compress(raw)

    def chunk(ctype, data):
        c = ctype + data
        crc = struct.pack('>I', zlib.crc32(c) & 0xffffffff)
        return struct.pack('>I', len(data)) + c + crc

    png = b'\x89PNG\r\n\x1a\n'
    png += chunk(b'IHDR', struct.pack('>IIBBBBB', w, h, 8, 2, 0, 0, 0))
    png += chunk(b'IDAT', compressed)
    png += chunk(b'IEND', b'')

    path = os.path.join(DATA_DIR, "test_image.png")
    with open(path, 'wb') as f:
        f.write(png)
    print("  test_image.png")


if __name__ == "__main__":
    print("Generating test fixtures...")
    gen_format_e()
    gen_format_d()
    gen_format_c()
    gen_format_b()
    gen_format_a()
    gen_empty()
    gen_multi_sheet()
    gen_format_e_regime()
    gen_test_image()
    print("Done!")
